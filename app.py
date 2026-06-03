import streamlit as st
import pandas as pd
import json
import requests
import os
import tempfile
import ssl

# ── SSL Bypass (Corporate Proxy) ─────────────────────────────────────────────
# Must be done BEFORE importing llama_parse or any llama-cloud libraries
# so that all httpx/requests connections skip certificate verification.
os.environ["PYTHONHTTPSVERIFY"] = "0"
os.environ["CURL_CA_BUNDLE"] = ""
os.environ["REQUESTS_CA_BUNDLE"] = ""
ssl._create_default_https_context = ssl._create_unverified_context

import httpx
# ─────────────────────────────────────────────────────────────────────────────

from io import BytesIO
from dotenv import load_dotenv
from llama_parse import LlamaParse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

# App configuration
st.set_page_config(
    page_title="AI Invoice Extraction Agent",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium styling
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Outfit', sans-serif;
    }
    
    .main-title {
        font-size: 3rem;
        font-weight: 700;
        background: linear-gradient(90deg, #FF4B4B, #FF8F8F, #4A90E2);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
    }
    
    .subtitle {
        font-size: 1.2rem;
        color: #6c757d;
        margin-bottom: 2rem;
    }
    
    .metric-card {
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 12px;
        padding: 1.5rem;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        transition: transform 0.2s, box-shadow 0.2s;
    }
    
    .metric-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 16px rgba(0, 0, 0, 0.2);
        border-color: rgba(255, 75, 75, 0.4);
    }
    
    .metric-val {
        font-size: 2.2rem;
        font-weight: 700;
        color: #FF4B4B;
    }
    
    .metric-lbl {
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 1px;
        color: #a1a1a1;
        margin-top: 0.5rem;
    }
    
    .success-badge {
        background-color: #d1fae5;
        color: #065f46;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.8rem;
    }
    
    .fail-badge {
        background-color: #fee2e2;
        color: #991b1b;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.8rem;
    }
    </style>
""", unsafe_allow_html=True)

# Sidebar Configuration
with st.sidebar:
    st.image("https://cdn-icons-png.flaticon.com/512/2822/2822839.png", width=80)
    st.markdown("## **Invoice Agent Config**")
    st.markdown("Configure the keys below. The application will first look for them in your `.env` file.")
    
    llama_key = st.text_input(
        "LlamaCloud API Key", 
        type="password",
        value=os.environ.get("LLAMA_CLOUD_API_KEY", ""),
        help="Get a key from https://cloud.llamaindex.ai/"
    )
    
    groq_key = st.text_input(
        "Groq API Key",
        type="password",
        value=os.environ.get("GROQ_API_KEY", ""),
        help="Get a key from https://console.groq.com/"
    )
    
    st.markdown("---")
    st.markdown("### **Model Selection**")
    MODELS = {
        "Llama 3.3 70B (Versatile)": "llama-3.3-70b-versatile",
        "Llama 3 8B (8192)": "llama3-8b-8192",
        "Mixtral 8x7B (32768)": "mixtral-8x7b-32768"
    }
    selected_model_name = st.selectbox("LLM Model for Extraction", list(MODELS.keys()), index=0)
    selected_model_id = MODELS[selected_model_name]
    
    st.markdown("---")
    st.markdown(
        "### **About LlamaParse**\n"
        "LlamaParse is a specialized document parser optimized for RAG. "
        "It excels at reading tables, complex formatting, and multi-column "
        "documents from PDFs and images."
    )

# Header Section
st.markdown("<h1 class='main-title'>📄 AI Invoice Extraction Agent</h1>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>Upload one or more invoices (PDF, PNG, JPG). The agent will parse them with LlamaParse and extract key structured information using LLMs.</p>", unsafe_allow_html=True)

# Main UI Columns
col_upload, col_info = st.columns([2, 1])

with col_info:
    st.info(
        "💡 **Expected Output Fields (1 row per invoice):**\n"
        "- **Supplier Name** (e.g. Acme Corp Ltd)\n"
        "- **Invoice No** (e.g. INV-12345)\n"
        "- **Invoice Date** (e.g. 2026-05-25)\n"
        "- **Part No** — all part numbers from the invoice (comma-separated)\n"
        "- **Total Quantity** — the combined total from the TOTAL row\n\n"
        "If a PDF contains 2 invoices on separate pages, each will produce its own row."
    )

with col_upload:
    uploaded_files = st.file_uploader(
        "Upload Invoices (PDF or Images)",
        type=["pdf", "png", "jpg", "jpeg"],
        accept_multiple_files=True,
        help="Select one or multiple files to process."
    )

# Session State Initializations
if "extracted_data" not in st.session_state:
    st.session_state.extracted_data = []
if "parsed_markdowns" not in st.session_state:
    st.session_state.parsed_markdowns = {}
if "processing_log" not in st.session_state:
    st.session_state.processing_log = []

def clean_json_string(text):
    text = text.strip()
    if text.startswith("```json"):
        text = text[7:]
    elif text.startswith("```"):
        text = text[3:]
    if text.endswith("```"):
        text = text[:-3]
    return text.strip()

def parse_invoice_with_llama(file_bytes, file_name, api_key):
    """
    Saves the file to a temp path and parses it using LlamaParse SDK
    """
    suffix = os.path.splitext(file_name)[1]
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        temp_file.write(file_bytes)
        temp_path = temp_file.name
    finally:
        temp_file.close()
        
    try:
        instruction = (
            "This document may contain one or more invoices, possibly on separate pages. "
            "Treat each page's invoice as a SEPARATE invoice. For EACH invoice found, extract ONLY: "
            "1. Supplier Name (supplier_name) — the company or vendor who issued the invoice (usually the letterhead at the top).\n"
            "2. Invoice Number (invoice_no) — the field explicitly labelled 'Invoice No', 'Invoice Number', or 'Invoice #'. "
            "DO NOT use Packing List No, Packing Slip No, Packaging List No, PL No, or any similar packing/packaging reference numbers. "
            "These are completely different documents and their numbers must be ignored.\n"
            "3. Invoice Date (invoice_date) — the date associated with the invoice, not the packing list.\n"
            "4. All Part Numbers (part_nos) from that invoice's line items, combined as a single comma-separated list.\n"
            "5. Total Quantity (total_quantity) — the single combined total quantity from the TOTAL row at the bottom of the invoice. "
            "Do NOT list individual quantities per line item; only the TOTAL row value is needed.\n"
            "Ignore all other details such as terms, billing/shipping addresses, tax rates, bank details. "
            "Format clearly as tables or lists, with a clear separator between invoices."
        )
        parser = LlamaParse(
            api_key=api_key,
            result_type="markdown",
            verbose=False,
            user_prompt=instruction
        )
        documents = parser.load_data(temp_path)
        markdown_text = "\n\n".join([doc.text for doc in documents])
        return markdown_text
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

def extract_details_from_markdown(markdown_content, model_id, api_key):
    """
    Calls Groq LLM to extract structured fields from parsed markdown
    """
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    prompt = f"""
    You are an expert data extraction agent. You will be provided with the markdown content of a parsed invoice document.
    This document may contain ONE or MORE separate invoices (often on different pages).
    Treat each page's invoice as a SEPARATE invoice.

    For EACH invoice found, extract:
    1. supplier_name — The company/vendor who issued the invoice (letterhead, "From", or top-of-page company name).
    2. invoice_no — The number labelled EXACTLY as "Invoice No", "Invoice Number", "Invoice #", or "Inv No".
       *** CRITICAL: Do NOT use Packing List No, Packing Slip No, Packaging List No, PL No, or any similar packing/packaging document numbers.
           These are completely different from invoice numbers. If you are unsure, prefer null over a wrong number. ***
    3. invoice_date — The date of the invoice (not a packing list date).
    4. part_nos — ALL part numbers from that invoice's line items combined into a single comma-separated string.
       Example: "21539393-21539393, 21702705-21702705"
       If no part numbers are found, set to null.
    5. total_quantity — The SINGLE combined total quantity shown in the TOTAL row at the bottom of the invoice.
       *** CRITICAL: Do NOT sum individual line item quantities yourself. Read the value from the row labelled "TOTAL" or "Total" in the invoice.
           If no TOTAL row exists, set to null. ***

    Invoice Markdown Content:
    ---
    {markdown_content}
    ---

    Return ONLY a JSON object with this exact structure (always use the "invoices" array even if there is only one invoice):
    {{
      "invoices": [
        {{
          "supplier_name": "str or null",
          "invoice_no": "str or null",
          "invoice_date": "str or null",
          "part_nos": "str or null",
          "total_quantity": "int or null"
        }}
      ]
    }}

    Rules:
    - If a field is not found, set its value to null.
    - If the document has 2 invoices on 2 different pages, return 2 objects inside the "invoices" array.
    - Each invoice object must represent ONE invoice completely independently.
    - Do NOT mix data from different invoices.
    - Do not add any explanation or conversational text. Return only valid JSON.
    """
    
    payload = {
        "model": model_id,
        "messages": [
            {"role": "system", "content": "You are a precise data extraction agent that outputs JSON only."},
            {"role": "user", "content": prompt}
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0
    }
    
    res = requests.post(url, headers=headers, json=payload, verify=False)
    res.raise_for_status()
    response_data = res.json()
    result_text = response_data['choices'][0]['message']['content']
    
    cleaned_json = clean_json_string(result_text)
    return json.loads(cleaned_json)

def generate_styled_excel(df):
    """
    Generates a professionally styled Excel sheet from the DataFrame using openpyxl
    """
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Invoice Data"
    
    # Write headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Style configuration
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Style header row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write and style data
    for row_idx, row in enumerate(df.values, start=2):
        ws.append(list(row))
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            
            # Alignments & formatting based on column
            col_name = headers[col_num - 1]
            if col_name in ["Supplier Name", "Invoice No", "Part No"]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_name == "Invoice Date":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "Total Quantity":
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output)
    output.seek(0)
    return output

# Trigger Extraction
if uploaded_files:
    if st.button("🚀 Process & Extract Invoices", type="primary"):
        if not llama_key:
            st.error("❌ LlamaCloud API Key is missing. Please provide it in the sidebar.")
            st.stop()
        if not groq_key:
            st.error("❌ Groq API Key is missing. Please provide it in the sidebar.")
            st.stop()
            
        st.session_state.extracted_data = []
        st.session_state.parsed_markdowns = {}
        st.session_state.processing_log = []
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_files = len(uploaded_files)
        
        for idx, uploaded_file in enumerate(uploaded_files):
            file_name = uploaded_file.name
            st.session_state.processing_log.append(f"Started processing {file_name}...")
            
            # Progress update
            pct = int((idx) / total_files * 100)
            progress_bar.progress(pct)
            status_text.markdown(f"**Processing ({idx+1}/{total_files}):** `{file_name}`...")
            
            try:
                # Step 1: LlamaParse
                status_text.markdown(f"**Step 1/2 Parsing with LlamaParse:** `{file_name}`...")
                file_bytes = uploaded_file.read()
                markdown_content = parse_invoice_with_llama(file_bytes, file_name, llama_key)
                
                # Save markdown for debugging/inspecting
                st.session_state.parsed_markdowns[file_name] = markdown_content
                
                # Step 2: Groq AI Extraction
                status_text.markdown(f"**Step 2/2 Extracting details with {selected_model_name}:** `{file_name}`...")
                extracted_json = extract_details_from_markdown(markdown_content, selected_model_id, groq_key)
                
                # Support both old single-invoice and new multi-invoice response shapes
                invoices = extracted_json.get("invoices", [])
                if not invoices:
                    # Fallback: handle old-style flat response gracefully
                    invoices = [{
                        "supplier_name": extracted_json.get("supplier_name"),
                        "invoice_no": extracted_json.get("invoice_no"),
                        "invoice_date": extracted_json.get("invoice_date"),
                        "part_nos": extracted_json.get("part_nos"),
                        "total_quantity": extracted_json.get("total_quantity")
                    }]

                for inv in invoices:
                    st.session_state.extracted_data.append({
                        "Source File": file_name,
                        "Supplier Name": inv.get("supplier_name"),
                        "Invoice No": inv.get("invoice_no"),
                        "Invoice Date": inv.get("invoice_date"),
                        "Part No": inv.get("part_nos") or "N/A",
                        "Total Quantity": inv.get("total_quantity") or 0,
                        "Status": "Success"
                    })
                
                st.session_state.processing_log.append(f"✅ Successfully processed {file_name}")
                
            except Exception as e:
                st.session_state.processing_log.append(f"❌ Error processing {file_name}: {str(e)}")
                st.session_state.extracted_data.append({
                    "Source File": file_name,
                    "Supplier Name": "Error",
                    "Invoice No": "Error",
                    "Invoice Date": "Error",
                    "Part No": "N/A",
                    "Total Quantity": 0,
                    "Status": f"Failed: {str(e)}"
                })
                
        progress_bar.progress(100)
        status_text.markdown("🎉 **All files processed!** Review the extracted details below.")

# Display Results
if st.session_state.extracted_data:
    df_results = pd.DataFrame(st.session_state.extracted_data)
    
    st.markdown("---")
    st.subheader("📊 Extraction Statistics & Metrics")
    
    # Calculate statistics
    success_count = df_results[df_results["Status"] == "Success"]["Source File"].nunique()
    total_invoices = df_results["Source File"].nunique()
    total_parts = len(df_results[df_results["Status"] == "Success"])
    total_qty = df_results["Total Quantity"].sum()
    
    m_col1, m_col2, m_col3, m_col4 = st.columns(4)
    with m_col1:
        st.markdown(
            f'<div class="metric-card">'
            f'<div class="metric-val">{total_invoices}</div>'
            f'<div class="metric-lbl">Total Uploaded</div>'
            f'</div>', 
            unsafe_allow_html=True
        )
    with m_col2:
        st.markdown(
            f'<div class="metric-card">'
            f'<div class="metric-val">{success_count}</div>'
            f'<div class="metric-lbl">Successfully Parsed</div>'
            f'</div>', 
            unsafe_allow_html=True
        )
    with m_col3:
        st.markdown(
            f'<div class="metric-card">'
            f'<div class="metric-val">{total_parts}</div>'
            f'<div class="metric-lbl">Invoices Extracted</div>'
            f'</div>', 
            unsafe_allow_html=True
        )
    with m_col4:
        st.markdown(
            f'<div class="metric-card">'
            f'<div class="metric-val">{int(total_qty):,}</div>'
            f'<div class="metric-lbl">Total Quantity</div>'
            f'</div>', 
            unsafe_allow_html=True
        )
        
    st.markdown("---")
    
    tab_data, tab_markdown, tab_logs = st.tabs([
        "📋 Extracted Excel Table", 
        "📄 LlamaParse Markdown View", 
        "⚙️ Processing Logs"
    ])
    
    with tab_data:
        st.markdown("### Extracted Invoice Data")
        st.markdown("Below is the tabular data extracted from the invoice(s).")
        
        # Display table with styled status
        def format_status(val):
            if val == "Success":
                return "color: green; font-weight: bold;"
            elif "Failed" in str(val):
                return "color: red; font-weight: bold;"
            else:
                return "color: orange; font-weight: bold;"
                
        styled_df = df_results.style.map(format_status, subset=["Status"])
        st.dataframe(styled_df, use_container_width=True)
        
        # Excel download button
        # Drop Status column for Excel download as per request to keep it clean, or keep it. Let's make it clean and exclude 'Source File' and 'Status' if they just want the columns 'Invoice No', 'Invoice Date', 'Part No', 'Part Quantity'.
        excel_df = df_results[df_results["Status"] != "Error"][["Supplier Name", "Invoice No", "Invoice Date", "Part No", "Total Quantity"]].copy()
        
        excel_data = generate_styled_excel(excel_df)
        
        st.download_button(
            label="📥 Download Extracted Data in Excel Format",
            data=excel_data,
            file_name="extracted_invoice_details.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
    with tab_markdown:
        st.markdown("### Raw Parsed Markdown")
        st.markdown("Select a file to inspect the raw Markdown text generated by LlamaParse. This helps verify layout mapping.")
        selected_file = st.selectbox("Select Invoice File", list(st.session_state.parsed_markdowns.keys()))
        if selected_file:
            st.markdown("---")
            st.markdown(st.session_state.parsed_markdowns[selected_file])
            
    with tab_logs:
        st.markdown("### Execution logs")
        for log in st.session_state.processing_log:
            if "❌" in log:
                st.error(log)
            elif "✅" in log:
                st.success(log)
            else:
                st.info(log)
else:
    st.markdown("---")
    st.info("ℹ️ Upload files and click **Process & Extract Invoices** to see results.")
